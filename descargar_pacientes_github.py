"""
Script para descargar reporte de pacientes atendidos de SUNUBE - Version GitHub Actions
Optimizado para ejecutarse en entorno headless de GitHub
Soporta multiples cuentas y combina los resultados
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
import time
import os
import logging
import glob
import shutil
import base64
import requests

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('descarga_pacientes.log'),
        logging.StreamHandler()
    ]
)

# Configuracion
URL_LOGIN = "https://hc.sunu.be/login"
URL_PACIENTES = "https://hc.sunu.be/reporte/pacientesAtendidosFecha"

# Lista de cuentas para descargar pacientes atendidos
CUENTAS = [
    {"nombre": "Daniel", "email": "drdanielavilaoftalmo@gmail.com", "password": "Gestion2025+"},
    {"nombre": "Carolina", "email": "campuzanocarolina14@gmail.com", "password": "43979299"},
]

# Carpeta de descargas (ruta absoluta requerida para CDP en headless)
DOWNLOAD_DIR = os.path.abspath(os.getcwd())

def configurar_chrome():
    """Configura opciones de Chrome para GitHub Actions (headless)"""
    chrome_options = Options()

    # Configuracion para headless (sin interfaz grafica)
    chrome_options.add_argument("--headless=new")  # Nueva sintaxis para Chrome 109+
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    # Configuracion de descargas
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    # Inicializar driver
    driver = webdriver.Chrome(options=chrome_options)

    # Habilitar descargas en modo headless via CDP
    driver.execute_cdp_cmd("Page.setDownloadBehavior", {
        "behavior": "allow",
        "downloadPath": DOWNLOAD_DIR
    })

    return driver

def hacer_login(driver, email, password, nombre_cuenta):
    """Realiza el login en la aplicacion"""
    logging.info(f"[{nombre_cuenta}] Navegando a pagina de login...")
    driver.get(URL_LOGIN)

    wait = WebDriverWait(driver, 15)

    try:
        email_field = wait.until(EC.presence_of_element_located((By.NAME, "email")))
        password_field = driver.find_element(By.NAME, "password")
    except:
        try:
            email_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email']")))
            password_field = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
        except:
            email_field = wait.until(EC.presence_of_element_located((By.ID, "email")))
            password_field = driver.find_element(By.ID, "password")

    logging.info(f"[{nombre_cuenta}] Ingresando credenciales...")
    email_field.clear()
    email_field.send_keys(email)

    password_field.clear()
    password_field.send_keys(password)

    try:
        login_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
    except:
        try:
            login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Ingresar') or contains(text(), 'Login') or contains(text(), 'Entrar')]")
        except:
            login_button = driver.find_element(By.TAG_NAME, "button")

    logging.info(f"[{nombre_cuenta}] Haciendo click en boton de login...")
    # Usar JavaScript click para mejor compatibilidad con headless
    driver.execute_script("arguments[0].click();", login_button)
    time.sleep(5)

    # Si el click no funciono, intentar submit del formulario
    if "login" in driver.current_url.lower():
        logging.info(f"[{nombre_cuenta}] Click no funciono, intentando submit del formulario...")
        driver.execute_script("document.querySelector('form').submit();")
        time.sleep(5)

    # Verificar que el login fue exitoso
    driver.save_screenshot(f"post_login_{nombre_cuenta}.png")
    logging.info(f"[{nombre_cuenta}] URL despues de login: {driver.current_url}")
    logging.info(f"[{nombre_cuenta}] Titulo despues de login: {driver.title}")

    # Verificar que no seguimos en la pagina de login
    if "login" in driver.current_url.lower():
        logging.error(f"[{nombre_cuenta}] Login parece haber fallado - aun en pagina de login")
        raise Exception(f"Login fallido para {nombre_cuenta} - aun en pagina de login")

    logging.info(f"[{nombre_cuenta}] Login completado!")

def descargar_reporte_pacientes(driver, nombre_cuenta):
    """Navega a la seccion de pacientes atendidos y descarga el reporte"""
    logging.info(f"[{nombre_cuenta}] Navegando a seccion de pacientes atendidos: {URL_PACIENTES}")
    driver.get(URL_PACIENTES)

    wait = WebDriverWait(driver, 20)

    # Verificar que la pagina cargo correctamente
    time.sleep(2)
    logging.info(f"[{nombre_cuenta}] URL actual: {driver.current_url}")
    logging.info(f"[{nombre_cuenta}] Titulo de pagina: {driver.title}")

    # Verificar si estamos en la pagina de login (sesion expirada)
    if "login" in driver.current_url.lower():
        logging.error(f"[{nombre_cuenta}] La sesion expiro o el login fallo - redirigido a login")
        driver.save_screenshot(f"sesion_expirada_{nombre_cuenta}.png")
        raise Exception(f"Sesion expirada para {nombre_cuenta} - redirigido a pagina de login")

    # Calcular fechas: ayer a ayer
    ayer = datetime.now() - timedelta(days=1)

    fecha_inicio_str = ayer.strftime("%Y-%m-%d")
    fecha_fin_str = ayer.strftime("%Y-%m-%d")

    logging.info(f"[{nombre_cuenta}] Rango de fechas: {fecha_inicio_str} a {fecha_fin_str}")

    time.sleep(3)

    # Tomar captura de pantalla para debugging
    driver.save_screenshot(f"pagina_pacientes_antes_{nombre_cuenta}.png")
    logging.info(f"[{nombre_cuenta}] Captura de pagina de pacientes guardada")

    # Log del HTML para debugging
    try:
        page_source = driver.page_source
        # Buscar inputs en la pagina
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')
        inputs = soup.find_all('input')
        logging.info(f"[{nombre_cuenta}] Inputs encontrados en la pagina: {len(inputs)}")
        for inp in inputs[:15]:  # Mostrar primeros 15 inputs
            logging.info(f"  Input: name='{inp.get('name', '')}' id='{inp.get('id', '')}' type='{inp.get('type', '')}' placeholder='{inp.get('placeholder', '')}'")
    except Exception as debug_e:
        logging.warning(f"[{nombre_cuenta}] No se pudo analizar HTML: {debug_e}")

    try:
        logging.info(f"[{nombre_cuenta}] Buscando campos de fecha...")

        # Intentar multiples selectores
        fecha_inicio_field = None
        fecha_fin_field = None

        # Opcion 1: Por ID (especifico de pacientes atendidos)
        try:
            fecha_inicio_field = wait.until(EC.presence_of_element_located((By.ID, "fecha_desde")))
            fecha_fin_field = driver.find_element(By.ID, "fecha_hasta")
            logging.info(f"[{nombre_cuenta}] Campos encontrados por ID (fecha_desde/fecha_hasta)")
        except Exception as e1:
            logging.warning(f"[{nombre_cuenta}] No se encontraron por ID: {e1}")

            # Opcion 2: Por name (desde/hasta)
            try:
                fecha_inicio_field = wait.until(EC.presence_of_element_located((By.NAME, "desde")))
                fecha_fin_field = driver.find_element(By.NAME, "hasta")
                logging.info(f"[{nombre_cuenta}] Campos encontrados por NAME (desde/hasta)")
            except Exception as e2:
                logging.warning(f"[{nombre_cuenta}] No se encontraron por NAME: {e2}")

            # Opcion 3: Por tipo date
            if not fecha_inicio_field:
                try:
                    date_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='date']")
                    logging.info(f"[{nombre_cuenta}] Inputs tipo date encontrados: {len(date_inputs)}")
                    if len(date_inputs) >= 2:
                        fecha_inicio_field = date_inputs[0]
                        fecha_fin_field = date_inputs[1]
                        logging.info(f"[{nombre_cuenta}] Campos encontrados por input[type='date']")
                except Exception as e3:
                    logging.warning(f"[{nombre_cuenta}] No se encontraron por type='date': {e3}")

        if not fecha_inicio_field or not fecha_fin_field:
            raise Exception(f"No se pudieron encontrar los campos de fecha con ningun selector para {nombre_cuenta}")

        logging.info(f"[{nombre_cuenta}] Llenando campos de fecha con JavaScript...")
        # Usar JavaScript para establecer valores (los campos son datepickers tipo text)
        driver.execute_script("arguments[0].value = arguments[1];", fecha_inicio_field, fecha_inicio_str)
        driver.execute_script("arguments[0].value = arguments[1];", fecha_fin_field, fecha_fin_str)

        time.sleep(1)

        logging.info(f"[{nombre_cuenta}] Buscando boton Enviar...")
        enviar_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Enviar')]")

        logging.info(f"[{nombre_cuenta}] Haciendo clic en boton Enviar...")
        enviar_button.click()

        logging.info(f"[{nombre_cuenta}] Esperando que carguen los resultados...")
        time.sleep(6)

        # Tomar captura despues de cargar resultados
        driver.save_screenshot(f"resultados_pacientes_{nombre_cuenta}.png")
        logging.info(f"[{nombre_cuenta}] Captura de resultados guardada")

        logging.info(f"[{nombre_cuenta}] Buscando boton de descarga/exportar...")
        download_button = None
        try:
            download_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Descargar') or contains(text(), 'Exportar') or contains(text(), 'Excel') or contains(text(), 'Export')]")
        except:
            try:
                download_button = driver.find_element(By.XPATH, "//a[contains(text(), 'Descargar') or contains(text(), 'Exportar') or contains(text(), 'Excel') or contains(text(), 'Export')]")
            except:
                try:
                    icon = driver.find_element(By.CSS_SELECTOR, "button i[class*='download'], a i[class*='download'], button i[class*='file'], a i[class*='file']")
                    download_button = icon.find_element(By.XPATH, "..")
                except:
                    pass

        # Diagnosticar el boton de descarga
        if download_button:
            button_html = driver.execute_script("return arguments[0].outerHTML;", download_button)
            button_tag = download_button.tag_name
            button_href = download_button.get_attribute('href')
            logging.info(f"[{nombre_cuenta}] Boton encontrado: tag={button_tag}, href={button_href}")
            logging.info(f"[{nombre_cuenta}] Boton HTML: {button_html[:500]}")

            # Verificar si esta dentro de un formulario
            form_info = driver.execute_script("""
                var form = arguments[0].closest('form');
                if (form) {
                    var inputs = {};
                    form.querySelectorAll('input').forEach(function(i) {
                        if (i.name) inputs[i.name] = i.value;
                    });
                    return {action: form.action, method: form.method || 'get', inputs: inputs};
                }
                return null;
            """, download_button)
            logging.info(f"[{nombre_cuenta}] Form info: {form_info}")
        else:
            button_href = None
            form_info = None
            logging.warning(f"[{nombre_cuenta}] No se encontro boton de descarga")

        # Inyectar interceptor de blobs antes de hacer clic
        driver.execute_script("""
            window.__capturedBlob = null;
            var origCreateObjectURL = URL.createObjectURL;
            URL.createObjectURL = function(blob) {
                var url = origCreateObjectURL.call(URL, blob);
                var reader = new FileReader();
                reader.readAsDataURL(blob);
                reader.onloadend = function() {
                    window.__capturedBlob = reader.result;
                };
                return url;
            };
        """)

        # Hacer clic en el boton de descarga
        if download_button:
            logging.info(f"[{nombre_cuenta}] Haciendo clic en boton de descarga...")
            driver.execute_script("arguments[0].click();", download_button)
            time.sleep(5)

        # === Estrategia 1: Verificar si Chrome descargo el archivo ===
        archivos = glob.glob(os.path.join(DOWNLOAD_DIR, "*.xlsx"))
        if archivos:
            archivos.sort(key=os.path.getmtime, reverse=True)
            archivo_descargado = archivos[0]
            nuevo_nombre = os.path.join(DOWNLOAD_DIR, f"reporte_pacientes_{nombre_cuenta}.xlsx")
            if os.path.exists(nuevo_nombre):
                os.remove(nuevo_nombre)
            shutil.move(archivo_descargado, nuevo_nombre)
            logging.info(f"[{nombre_cuenta}] Archivo descargado via Chrome: {nuevo_nombre}")
            return nuevo_nombre

        logging.info(f"[{nombre_cuenta}] Chrome no descargo archivo, intentando alternativas...")

        # === Estrategia 2: Verificar blob interceptado ===
        time.sleep(2)
        captured = driver.execute_script("return window.__capturedBlob;")
        if captured:
            logging.info(f"[{nombre_cuenta}] Blob interceptado! Guardando archivo...")
            data = base64.b64decode(captured.split(',')[1])
            filepath = os.path.join(DOWNLOAD_DIR, f"reporte_pacientes_{nombre_cuenta}.xlsx")
            with open(filepath, 'wb') as f:
                f.write(data)
            return filepath

        # === Estrategia 3: Descargar via requests con cookies de Selenium ===
        logging.info(f"[{nombre_cuenta}] Intentando descarga directa via requests...")
        cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        session = requests.Session()
        for name, value in cookies.items():
            session.cookies.set(name, value)
        session.headers.update({
            'User-Agent': driver.execute_script("return navigator.userAgent;")
        })

        # Intentar con href del boton o action del formulario
        download_url = None
        if button_href and button_href.startswith('http'):
            download_url = button_href
        elif form_info and form_info.get('action'):
            download_url = form_info['action']

        if download_url:
            logging.info(f"[{nombre_cuenta}] Descargando desde URL: {download_url}")
            if form_info and form_info.get('method', '').lower() == 'post':
                response = session.post(download_url, data=form_info.get('inputs', {}))
            else:
                response = session.get(download_url)

            if response.status_code == 200 and len(response.content) > 100:
                filepath = os.path.join(DOWNLOAD_DIR, f"reporte_pacientes_{nombre_cuenta}.xlsx")
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                logging.info(f"[{nombre_cuenta}] Archivo descargado via requests: {filepath} ({len(response.content)} bytes)")
                return filepath
            else:
                logging.warning(f"[{nombre_cuenta}] Requests fallo: status={response.status_code}, size={len(response.content)}")

        # === Estrategia 4: POST directo al endpoint de reporte ===
        logging.info(f"[{nombre_cuenta}] Intentando POST directo al endpoint...")
        csrf_token = driver.execute_script(
            "return document.querySelector('input[name=_token]')?.value || "
            "document.querySelector('meta[name=csrf-token]')?.content || ''")

        export_urls = [
            f"{URL_PACIENTES}/exportar",
            f"{URL_PACIENTES}/export",
            f"{URL_PACIENTES}/excel",
            URL_PACIENTES,
        ]
        for url in export_urls:
            try:
                response = session.post(url, data={
                    '_token': csrf_token,
                    'desde': fecha_inicio_str,
                    'hasta': fecha_fin_str,
                    'reporte': 'pacientesAtendidosFecha',
                })
                content_type = response.headers.get('content-type', '')
                logging.info(f"[{nombre_cuenta}] POST {url}: status={response.status_code}, content-type={content_type}, size={len(response.content)}")
                if response.status_code == 200 and ('spreadsheet' in content_type or 'excel' in content_type or 'octet-stream' in content_type):
                    filepath = os.path.join(DOWNLOAD_DIR, f"reporte_pacientes_{nombre_cuenta}.xlsx")
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    logging.info(f"[{nombre_cuenta}] Archivo descargado via POST directo: {filepath}")
                    return filepath
            except Exception as req_e:
                logging.warning(f"[{nombre_cuenta}] Error en POST a {url}: {req_e}")

        logging.warning(f"[{nombre_cuenta}] Todas las estrategias de descarga fallaron")
        todos = os.listdir(DOWNLOAD_DIR)
        logging.warning(f"[{nombre_cuenta}] Archivos en directorio: {[f for f in todos if not f.startswith('.')]}")
        return None

    except Exception as e:
        logging.error(f"[{nombre_cuenta}] Error al descargar reporte: {str(e)}")
        driver.save_screenshot(f"error_descarga_{nombre_cuenta}.png")
        raise

def combinar_excels(archivos_excel):
    """Combina multiples archivos Excel en uno solo"""
    from openpyxl import Workbook, load_workbook

    logging.info(f"Combinando {len(archivos_excel)} archivos Excel...")

    wb_combinado = Workbook()
    ws_combinado = wb_combinado.active
    ws_combinado.title = "Pacientes Combinados"

    fila_actual = 1
    encabezados_escritos = False

    for archivo in archivos_excel:
        if archivo is None:
            continue

        logging.info(f"Procesando: {archivo}")

        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                # Saltar encabezados en archivos posteriores al primero
                if idx == 0:
                    if encabezados_escritos:
                        continue
                    else:
                        encabezados_escritos = True

                for col_idx, valor in enumerate(row, 1):
                    ws_combinado.cell(row=fila_actual, column=col_idx, value=valor)
                fila_actual += 1

            wb.close()
            logging.info(f"Archivo procesado: {archivo}")

        except Exception as e:
            logging.error(f"Error procesando {archivo}: {e}")

    # Guardar archivo combinado
    archivo_combinado = os.path.join(DOWNLOAD_DIR, "reporte_pacientes_combinado.xlsx")
    wb_combinado.save(archivo_combinado)
    logging.info(f"Archivo combinado guardado: {archivo_combinado}")

    return archivo_combinado

def main():
    """Funcion principal"""
    logging.info("="*60)
    logging.info("DESCARGA DE REPORTE DE PACIENTES ATENDIDOS - SUNUBE (GitHub Actions)")
    logging.info(f"Procesando {len(CUENTAS)} cuentas...")
    logging.info("="*60)

    archivos_descargados = []
    errores = []

    for cuenta in CUENTAS:
        nombre = cuenta["nombre"]
        email = cuenta["email"]
        password = cuenta["password"]

        logging.info(f"\n{'='*60}")
        logging.info(f"PROCESANDO CUENTA: {nombre} ({email})")
        logging.info(f"{'='*60}")

        driver = None
        try:
            driver = configurar_chrome()
            hacer_login(driver, email, password, nombre)
            archivo = descargar_reporte_pacientes(driver, nombre)

            if archivo:
                archivos_descargados.append(archivo)
                logging.info(f"[{nombre}] Descarga completada exitosamente")
            else:
                errores.append(f"{nombre}: No se pudo descargar el archivo")

        except Exception as e:
            logging.error(f"[{nombre}] Error durante la ejecucion: {str(e)}")
            errores.append(f"{nombre}: {str(e)}")
            if driver:
                driver.save_screenshot(f"error_fatal_{nombre}.png")

        finally:
            if driver:
                driver.quit()
                logging.info(f"[{nombre}] Navegador cerrado")

    # Combinar archivos descargados
    logging.info(f"\n{'='*60}")
    logging.info("COMBINANDO ARCHIVOS")
    logging.info(f"{'='*60}")

    if archivos_descargados:
        try:
            archivo_final = combinar_excels(archivos_descargados)
            logging.info(f"Archivo final combinado: {archivo_final}")
        except Exception as e:
            logging.error(f"Error al combinar archivos: {e}")
            errores.append(f"Combinacion: {str(e)}")
    else:
        logging.error("No hay archivos para combinar")
        errores.append("No se descargo ningun archivo")

    # Resumen final
    logging.info(f"\n{'='*60}")
    logging.info("RESUMEN FINAL")
    logging.info(f"{'='*60}")
    logging.info(f"Cuentas procesadas: {len(CUENTAS)}")
    logging.info(f"Archivos descargados: {len(archivos_descargados)}")

    if errores:
        logging.error(f"Errores encontrados ({len(errores)}):")
        for error in errores:
            logging.error(f"  - {error}")
        exit(1)
    else:
        logging.info("PROCESO COMPLETADO EXITOSAMENTE")
        logging.info(f"{'='*60}")

if __name__ == "__main__":
    main()
